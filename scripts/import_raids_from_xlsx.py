import argparse
import json
import math
import re
import sys
import time
import warnings
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Run: python -m pip install -r scripts/requirements.txt")
    sys.exit(1)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RAIDS_PATH = ROOT / "data" / "raids.json"
DEFAULT_SHEET_NAME = "Serca+Cath"

EXACT_COLOR_NAMES = {
    (0, 255, 255): "Cyan",
    (52, 168, 83): "Green",
    (66, 133, 244): "Blue",
    (103, 78, 167): "Purple",
    (127, 96, 0): "Brown",
    (133, 32, 12): "Brick Red",
    (147, 196, 125): "Light Green",
    (153, 0, 255): "Purple",
    (153, 153, 153): "Gray",
    (163, 232, 85): "Lime",
    (201, 218, 248): "Light Blue",
    (224, 102, 102): "Red",
    (244, 199, 195): "Pink",
    (251, 188, 4): "Gold",
    (255, 0, 255): "Magenta",
    (255, 153, 0): "Orange",
    (255, 229, 153): "Light Yellow",
}

COLOR_PALETTE = {
    "Red": (224, 92, 96),
    "Orange": (255, 153, 0),
    "Amber": (251, 188, 4),
    "Gold": (247, 203, 77),
    "Light Yellow": (255, 229, 153),
    "Lime": (163, 232, 85),
    "Green": (100, 180, 90),
    "Light Green": (147, 196, 125),
    "Cyan": (0, 210, 220),
    "Light Blue": (201, 218, 248),
    "Blue": (66, 133, 244),
    "Purple": (120, 80, 170),
    "Pink": (238, 90, 185),
    "Magenta": (255, 0, 255),
    "Brown": (140, 50, 25),
    "Gray": (150, 150, 150),
}

KNOWN_RAIDS = ("Serca", "Cathedral")
KNOWN_DIFFICULTIES = ("Nightmare", "Hard", "Normal", "2", "3")
THEME_RGBS = []


def normalize_player_name(name):
    return name.strip().split("-")[0].strip().lower()


def clean_text(value):
    if value is None:
        return ""

    return re.sub(r"\s+", " ", str(value)).strip()


def rgb_from_color(color):
    if color.type == "rgb" and color.rgb:
        value = color.rgb[-6:]
        try:
            return tuple(int(value[index : index + 2], 16) for index in (0, 2, 4))
        except ValueError:
            return None

    if color.type == "theme" and color.theme is not None and color.theme < len(THEME_RGBS):
        return THEME_RGBS[color.theme]

    return None


def rgb_from_cell(cell):
    if cell.fill.fill_type is None:
        return None

    color = cell.fill.fgColor
    return rgb_from_color(color)


def load_theme_rgbs(workbook):
    global THEME_RGBS

    theme_xml = workbook.loaded_theme
    if not theme_xml:
        THEME_RGBS = []
        return

    namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    root = ET.fromstring(theme_xml)
    color_scheme = root.find(".//a:clrScheme", namespace)

    if color_scheme is None:
        THEME_RGBS = []
        return

    colors = []
    for child in list(color_scheme):
        srgb = child.find(".//a:srgbClr", namespace)
        system = child.find(".//a:sysClr", namespace)
        value = srgb.attrib.get("val") if srgb is not None else None
        value = value or (system.attrib.get("lastClr") if system is not None else None)

        if value:
            colors.append(tuple(int(value[index : index + 2], 16) for index in (0, 2, 4)))

    THEME_RGBS = colors


def nearest_color_name(rgb):
    if rgb is None:
        return "Unknown"

    if rgb in EXACT_COLOR_NAMES:
        return EXACT_COLOR_NAMES[rgb]

    best_name = "Unknown"
    best_distance = math.inf

    for name, palette_rgb in COLOR_PALETTE.items():
        distance = sum((rgb[index] - palette_rgb[index]) ** 2 for index in range(3))
        if distance < best_distance:
            best_name = name
            best_distance = distance

    return best_name


def parse_raid_label(value):
    text = clean_text(value)
    match = re.match(r"^(Serca|Cathedral)\b\s*(.*)$", text, flags=re.IGNORECASE)

    if not match:
        return None, None

    raid_name = next(raid for raid in KNOWN_RAIDS if raid.lower() == match.group(1).lower())
    difficulty = match.group(2).strip()

    if difficulty not in KNOWN_DIFFICULTIES:
        return None, None

    return raid_name, difficulty


def is_raid_label(value):
    raid_name, _ = parse_raid_label(value)
    return raid_name is not None


def classify_role(cell):
    rgb = rgb_from_cell(cell)

    if rgb is None:
        return "DPS"

    _, g, b = rgb
    return "Support" if b - g > 8 else "DPS"


def parse_member(value):
    text = clean_text(value)

    if not text:
        return None

    name, separator, label = text.partition("-")

    member = {
        "name": name.strip(),
        "lookupName": normalize_player_name(name),
    }

    if separator and label.strip():
        member["label"] = label.strip()

    return member


def parse_item_level(value):
    if value is None:
        return None

    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def parse_raid_status(sheet, row, col):
    status_cells = [
        sheet.cell(row=row + offset, column=col).value
        for offset in range(1, 4)
        if row + offset <= sheet.max_row
    ]

    if any(clean_text(value).upper() == "DONE" for value in status_cells):
        return "DONE"

    return "TODO"


def parse_raid_block(sheet, formula_sheet, row, col):
    raid_name, difficulty = parse_raid_label(sheet.cell(row=row, column=col).value)

    if not raid_name:
        return None

    members = []
    next_label_col = sheet.max_column + 1

    for scan_col in range(col + 1, sheet.max_column + 1):
        if is_raid_label(sheet.cell(row=row, column=scan_col).value):
            next_label_col = scan_col
            break

    for member_col in range(col + 1, next_label_col):
        # Rows beneath a raid contain formulas whose cached values are class names
        # (for example, Shadowhunter or Bard). They are raid details, not players.
        if formula_sheet.cell(row=row, column=member_col).data_type == "f":
            continue

        member_value = clean_text(sheet.cell(row=row, column=member_col).value)

        # Player entries use the "player-character" format. Requiring that
        # separator also prevents adjacent detail labels such as Serca1,
        # Serca2, and Reclear from turning a detail row into a raid.
        if "-" not in member_value:
            continue

        member = parse_member(member_value)

        if not member:
            continue

        role_cell = sheet.cell(row=row + 1, column=member_col)
        item_level = parse_item_level(sheet.cell(row=row + 2, column=member_col).value)
        member["role"] = classify_role(role_cell)
        if item_level is not None:
            member["itemLevel"] = item_level
        members.append(member)

        if len(members) == 4:
            break

    if not members:
        return None

    return {
        "id": "",
        "color": nearest_color_name(rgb_from_cell(sheet.cell(row=row, column=col))),
        "name": raid_name,
        "difficulty": difficulty,
        "status": parse_raid_status(sheet, row, col),
        "members": members,
        "createdBy": "xlsx-import",
        "createdAt": "",
        "_sourceCell": sheet.cell(row=row, column=col).coordinate,
    }


def parse_workbook(workbook_path, sheet_name, debug=False):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(workbook_path, data_only=True)
        formula_workbook = load_workbook(workbook_path, data_only=False)

    load_theme_rgbs(workbook)

    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        print(f"Sheet not found: {sheet_name}. Available sheets: {available}")
        sys.exit(1)

    sheet = workbook[sheet_name]
    formula_sheet = formula_workbook[sheet_name]
    raids = []

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row, column=col).value

            if not is_raid_label(value):
                continue

            raid = parse_raid_block(sheet, formula_sheet, row, col)

            if raid:
                raids.append(raid)
                if debug:
                    print(f"Parsed {raid['_sourceCell']}: {raid['color']} {raid['name']} {raid['difficulty']}")
            elif debug:
                print(f"Skipped {sheet.cell(row=row, column=col).coordinate}: no members parsed")

    now_ms = int(time.time() * 1000)
    created_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    for index, raid in enumerate(raids):
        raid["id"] = str(now_ms + index)
        raid["createdAt"] = created_at
        del raid["_sourceCell"]

    return raids


def preview_raids(raids):
    print()
    print(f"Parsed {len(raids)} raid(s):")
    print()

    for index, raid in enumerate(raids, start=1):
        members = ", ".join(f"{member['name']} ({member['role']})" for member in raid["members"])
        print(f"{index:>2}. [{raid['status']}] {raid['color']} {raid['name']} {raid['difficulty']} - {members}")

    print()


def write_raids(raids_path, raids):
    raids_path.parent.mkdir(parents=True, exist_ok=True)
    raids_path.write_text(json.dumps(raids, indent=2) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(
        description="Import fixed raid groups from the Copy of Serca+Cath workbook sheet into data/raids.json."
    )
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx workbook")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Sheet to import")
    parser.add_argument("--output", type=Path, default=DEFAULT_RAIDS_PATH, help="Path to raids.json")
    parser.add_argument("--debug", action="store_true", help="Print parser diagnostics")
    parser.add_argument("--json", action="store_true", help="Print parsed raids as JSON and exit")
    parser.add_argument("--yes", action="store_true", help="Replace raids without an interactive confirmation")
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}")
        sys.exit(1)

    raids = parse_workbook(args.workbook, args.sheet, debug=args.debug)

    if not raids:
        print("No raids were parsed. Nothing was changed.")
        print(f"Make sure the workbook has a sheet named {args.sheet} and visible raid labels like Serca Hard.")
        sys.exit(1)

    if args.json:
        print(json.dumps(raids))
        return

    preview_raids(raids)

    if args.yes:
        write_raids(args.output, raids)
        print(f"Imported {len(raids)} raid(s) into {args.output}.")
        return

    try:
        approval = input(f"Replace all raids in {args.output} with these results? Type YES to continue: ")
    except EOFError:
        print("Import cancelled. No confirmation was provided.")
        return

    if approval != "YES":
        print("Import cancelled. Nothing was changed.")
        return

    write_raids(args.output, raids)
    print(f"Imported {len(raids)} raid(s) into {args.output}.")


if __name__ == "__main__":
    main()
