import argparse
import copy
import json
import math
import random
import re
import sys
import time
import warnings
from pathlib import Path

from import_raids_from_xlsx import (
    DEFAULT_SHEET_NAME,
    classify_role,
    clean_text,
    is_raid_label,
    load_theme_rgbs,
    nearest_color_name,
    normalize_player_name,
    parse_raid_label,
    parse_raid_status,
    rgb_from_cell,
)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Run: python -m pip install -r scripts/requirements.txt")
    sys.exit(1)


ROOT = Path(__file__).resolve().parents[1]
COLOR_POOL = [
    "Red",
    "Orange",
    "Amber",
    "Gold",
    "Light Yellow",
    "Lime",
    "Green",
    "Light Green",
    "Cyan",
    "Light Blue",
    "Purple",
    "Pink",
    "Magenta",
    "Brown",
    "Gray",
    "Brick Red",
]


def parse_member(value):
    text = clean_text(value)
    if not text:
        return None

    name, separator, label = text.partition("-")
    member = {
        "name": name.strip(),
        "lookupName": normalize_player_name(name),
    }
    if separator and label.strip():
        member["label"] = label.strip()
    return member


def parse_item_level(value):
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def parse_workbook(workbook_path, sheet_name):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(workbook_path, data_only=True)

    load_theme_rgbs(workbook)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    raids = []

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            raid_name, difficulty = parse_raid_label(sheet.cell(row=row, column=col).value)
            if not raid_name:
                continue

            members = []
            next_label_col = sheet.max_column + 1
            for scan_col in range(col + 1, sheet.max_column + 1):
                if is_raid_label(sheet.cell(row=row, column=scan_col).value):
                    next_label_col = scan_col
                    break

            for member_col in range(col + 1, next_label_col):
                member = parse_member(sheet.cell(row=row, column=member_col).value)
                if not member:
                    continue

                item_level = parse_item_level(sheet.cell(row=row + 2, column=member_col).value)
                member["role"] = classify_role(sheet.cell(row=row + 1, column=member_col))
                if item_level is not None:
                    member["itemLevel"] = item_level
                members.append(member)

                if len(members) == 4:
                    break

            if not members:
                continue

            raids.append(
                {
                    "color": nearest_color_name(rgb_from_cell(sheet.cell(row=row, column=col))),
                    "name": raid_name,
                    "difficulty": difficulty,
                    "status": parse_raid_status(sheet, row, col),
                    "members": members,
                    "originalIndex": len(raids),
                }
            )

    return raids


def canonical_members(raid):
    return "|".join(sorted(member["lookupName"] for member in raid["members"]))


def member_key(member, raid):
    return "|".join(
        [
            raid["name"],
            member["role"],
            member["name"],
            member.get("lookupName", ""),
            member.get("label", ""),
            str(member.get("itemLevel", "")),
        ]
    )


def build_state(raids):
    state = copy.deepcopy(raids)
    for raid in state:
        raid["locked"] = raid["name"] == "Serca" and raid["difficulty"] == "Nightmare"
        raid["originalColor"] = raid["color"]
        raid["originalMemberKeys"] = [member_key(member, raid) for member in raid["members"]]
        for member in raid["members"]:
            if member.get("itemLevel") is None:
                member["eligibleForCathedral3"] = raid["name"] == "Cathedral" and raid["difficulty"] == "3"
            else:
                member["eligibleForCathedral3"] = member["itemLevel"] >= 1750
            member["originalMemberKey"] = member_key(member, raid)
    return state


def repair_mixed_colors(raids, color_pool):
    used_colors = set(raid["color"] for raid in raids)

    for cluster in list(clusters_for(raids)):
        groups = {}
        for raid in cluster:
            groups.setdefault(canonical_members(raid), []).append(raid)

        if len(groups) <= 1:
            continue

        keep_key = None
        for key, group in groups.items():
            if any(raid["locked"] for raid in group):
                keep_key = key
                break
        keep_key = keep_key or max(groups, key=lambda key: len(groups[key]))

        for key, group in groups.items():
            if key == keep_key:
                continue

            available = [
                color for color in color_pool
                if color not in used_colors or all(
                    canonical_members(raid) == key for raid in raids if raid["color"] == color
                )
            ]
            if not available:
                continue

            new_color = available[0]
            used_colors.add(new_color)
            for raid in group:
                raid["color"] = new_color


def validate_raid(raid):
    problems = []
    dps = sum(1 for member in raid["members"] if member["role"] == "DPS")
    supports = sum(1 for member in raid["members"] if member["role"] == "Support")
    names = [member["lookupName"] for member in raid["members"]]

    if dps > 3:
        problems.append("more than 3 DPS")
    if supports > 1:
        problems.append("more than 1 Support")
    if len(raid["members"]) > 4:
        problems.append("more than 4 members")
    if len(names) != len(set(names)):
        problems.append("duplicate player")
    if raid["name"] == "Cathedral" and raid["difficulty"] == "3":
        bad = [member["name"] for member in raid["members"] if not member.get("eligibleForCathedral3")]
        if bad:
            problems.append(f"Cathedral 3 has non-1750 member(s): {', '.join(bad)}")
    if raid["name"] == "Cathedral" and raid["difficulty"] == "2":
        bad = [member["name"] for member in raid["members"] if member.get("eligibleForCathedral3")]
        if bad:
            problems.append(f"Cathedral 2 has 1750 member(s): {', '.join(bad)}")
    return problems


def clusters_for(raids):
    clusters = {}
    for raid in raids:
        clusters.setdefault(raid["color"], []).append(raid)
    return sorted(clusters.values(), key=lambda cluster: (-len(cluster), cluster[0]["color"]))


def cluster_size_score(size):
    if size == 1:
        return -70
    if size in (2, 5):
        return 80
    if size == 3:
        return 190
    if size == 4:
        return 260
    return -260 * (size - 5)


def validate_cluster(cluster):
    if len(cluster) <= 1:
        return []
    expected = canonical_members(cluster[0])
    if all(canonical_members(raid) == expected for raid in cluster):
        return []
    return ["color mixes different player groups; every run in a color must have the exact same players"]


def score_state(raids):
    score = 0
    problems = []

    for raid in raids:
        dps = sum(1 for member in raid["members"] if member["role"] == "DPS")
        supports = sum(1 for member in raid["members"] if member["role"] == "Support")
        if dps == 3 and supports == 1:
            score += 12
        elif dps == 2 and supports == 1 and len(raid["members"]) == 3:
            score -= 8
        else:
            score -= 35

        raid_problems = validate_raid(raid)
        if raid_problems:
            problems.append((f"{raid['color']} {raid['name']} {raid['difficulty']}", raid_problems))
            score -= 1000 * len(raid_problems)

    for cluster in clusters_for(raids):
        cluster_problems = validate_cluster(cluster)
        if cluster_problems:
            problems.append((f"{cluster[0]['color']} Color cluster", cluster_problems))
            score -= 3000 * len(cluster_problems)
            continue
        score += cluster_size_score(len(cluster))
        if len(cluster) <= 5:
            score += len(cluster) * 80

    return score, problems


def can_swap(raids, left, right):
    left_raid = raids[left[0]]
    right_raid = raids[right[0]]
    if left[0] == right[0]:
        return False
    if left_raid["name"] != right_raid["name"]:
        return False

    left_member = left_raid["members"][left[1]]
    right_member = right_raid["members"][right[1]]
    if left_member["role"] != right_member["role"]:
        return False

    if left_raid["name"] == "Cathedral" and left_raid["difficulty"] == "3" and not right_member.get("eligibleForCathedral3"):
        return False
    if right_raid["name"] == "Cathedral" and right_raid["difficulty"] == "3" and not left_member.get("eligibleForCathedral3"):
        return False
    if left_raid["name"] == "Cathedral" and left_raid["difficulty"] == "2" and right_member.get("eligibleForCathedral3"):
        return False
    if right_raid["name"] == "Cathedral" and right_raid["difficulty"] == "2" and left_member.get("eligibleForCathedral3"):
        return False
    return True


def allowed_colors(raids, raid_index, color_pool):
    members = canonical_members(raids[raid_index])
    allowed = []
    for color in color_pool:
        if all(index == raid_index or raid["color"] != color or canonical_members(raid) == members for index, raid in enumerate(raids)):
            allowed.append(color)
    return allowed


def signature(raids):
    return ";".join(
        f"{raid['originalIndex']}:{raid['color']}:{canonical_members(raid)}" for raid in raids
    )


def changed_counts(raids):
    changed = 0
    color_changed = 0
    for raid in raids:
        raid_changed = raid["color"] != raid["originalColor"] or any(
            member["originalMemberKey"] != raid["originalMemberKeys"][index]
            for index, member in enumerate(raid["members"])
        )
        if raid_changed:
            changed += 1
        if raid["color"] != raid["originalColor"]:
            color_changed += 1
    return changed, color_changed


def adjusted_score(score, changed, color_changed, variety):
    return score + changed * variety * 8 + color_changed * variety * 14


def optimize_once(source_raids, iterations, variety):
    raids = build_state(source_raids)
    color_pool = sorted(set([raid["color"] for raid in raids] + COLOR_POOL))
    repair_mixed_colors(raids, color_pool)
    slots = [
        (raid_index, member_index)
        for raid_index, raid in enumerate(raids)
        if not raid["locked"]
        for member_index, _ in enumerate(raid["members"])
    ]
    recolor_slots = list(range(len(raids)))
    current_score, _ = score_state(raids)
    initial = signature(raids)
    best = None
    best_adjusted = -math.inf
    temperature = 28 + variety * 4

    for _ in range(iterations):
        undo = None
        if random.random() < 0.35:
            raid_index = random.choice(recolor_slots)
            colors = allowed_colors(raids, raid_index, color_pool)
            if not colors:
                continue
            color = random.choice(colors)
            if raids[raid_index]["color"] == color:
                continue
            previous = raids[raid_index]["color"]
            raids[raid_index]["color"] = color
            undo = lambda ri=raid_index, pc=previous: raids[ri].__setitem__("color", pc)
        else:
            left = random.choice(slots)
            right = random.choice(slots)
            if not can_swap(raids, left, right):
                continue
            raids[left[0]]["members"][left[1]], raids[right[0]]["members"][right[1]] = (
                raids[right[0]]["members"][right[1]],
                raids[left[0]]["members"][left[1]],
            )
            def undo_swap(l=left, r=right):
                raids[l[0]]["members"][l[1]], raids[r[0]]["members"][r[1]] = (
                    raids[r[0]]["members"][r[1]],
                    raids[l[0]]["members"][l[1]],
                )
            undo = undo_swap

        next_score, problems = score_state(raids)
        changed, color_changed = changed_counts(raids)
        if not problems and changed and signature(raids) != initial:
            candidate_adjusted = adjusted_score(next_score, changed, color_changed, variety)
            if candidate_adjusted > best_adjusted:
                best_adjusted = candidate_adjusted
                best = copy.deepcopy(raids)

        delta = next_score - current_score
        if delta >= 0 or math.exp(delta / temperature) > random.random():
            current_score = next_score
        else:
            undo()

        temperature = max(0.35, temperature * 0.99992)

    if best is None:
        best = raids
    score, problems = score_state(best)
    changed, color_changed = changed_counts(best)
    return {
        "score": score,
        "problems": problems,
        "changed": changed,
        "colorChanged": color_changed,
        "clusters": clusters_for(best),
        "raids": best,
    }


def format_member(member):
    label = f"-{member['label']}" if member.get("label") else ""
    item_level = f" {member['itemLevel']}" if member.get("itemLevel") is not None else ""
    return f"{member['name']}{label} ({member['role']}{item_level})"


def format_report(baseline, suggestions):
    lines = [
        "Raid group suggestions",
        "======================",
        "",
        f"Current score: {baseline[0]}",
        "Current validation issues: None" if not baseline[1] else f"Current validation issues: {baseline[1]}",
        "",
    ]
    for option_index, suggestion in enumerate(suggestions, start=1):
        delta = suggestion["score"] - baseline[0]
        lines.extend(
            [
                f"Option {option_index}",
                "--------",
                f"Score: {suggestion['score']} ({delta:+d})",
                f"Changed raids: {suggestion['changed']}",
                f"Color changes: {suggestion['colorChanged']}",
                "Validation issues: None",
                "",
                "Clusters:",
            ]
        )
        for cluster in suggestion["clusters"]:
            members = ", ".join(sorted(member["name"] for member in cluster[0]["members"]))
            raids = "; ".join(
                f"{raid['name']} {raid['difficulty']} [{', '.join(sorted(member['name'] for member in raid['members']))}]"
                for raid in cluster
            )
            lines.append(f"{cluster[0]['color']}: {len(cluster)} run(s), shared [{members}] -> {raids}")
        lines.extend(["", "Groups:"])
        for raid in sorted(suggestion["raids"], key=lambda raid: raid["originalIndex"]):
            lines.append(
                f"{raid['color']} {raid['name']} {raid['difficulty']}: "
                + ", ".join(format_member(member) for member in raid["members"])
            )
        lines.append("")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--options", type=int, default=3)
    parser.add_argument("--search", type=int, default=3)
    parser.add_argument("--variety", type=int, default=5)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    random.seed(time.time())
    raids = parse_workbook(args.workbook, args.sheet)
    baseline = score_state(build_state(raids))
    suggestions = []
    seen = set()
    attempts = max(args.options * 5, 8)
    iterations = args.search * 20000

    for attempt in range(attempts):
        if len(suggestions) >= args.options:
            break
        suggestion = optimize_once(raids, iterations, args.variety + attempt)
        sig = signature(suggestion["raids"])
        if sig in seen or suggestion["problems"] or suggestion["changed"] == 0:
            continue
        seen.add(sig)
        suggestions.append(suggestion)

    suggestions.sort(
        key=lambda suggestion: adjusted_score(
            suggestion["score"], suggestion["changed"], suggestion["colorChanged"], args.variety
        ),
        reverse=True,
    )
    report = format_report(baseline, suggestions[: args.options])
    if args.output:
        args.output.write_text(report + "\n", encoding="utf-8")
    print(report)


if __name__ == "__main__":
    main()
