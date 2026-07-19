import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    print("Missing dependency. Run: python -m pip install -r scripts/requirements.txt")
    sys.exit(1)

from render_schedule_from_xlsx import render_schedule


COLOR_HEX = {
    "Red": "E06666",
    "Orange": "FF9900",
    "Amber": "F9AB00",
    "Gold": "FBBC04",
    "Light Yellow": "FFE599",
    "Lime": "A3E855",
    "Green": "34A853",
    "Light Green": "93C47D",
    "Cyan": "00FFFF",
    "Light Blue": "C9DAF8",
    "Purple": "674EA7",
    "Pink": "F4C7C3",
    "Magenta": "FF00FF",
    "Brown": "85200C",
    "Brick Red": "85200C",
    "Gray": "999999",
}

BLACK = "000000"
HEADER_BLUE = "9FC5E8"
DPS_GREEN = "D9EAD3"
SUPPORT_PURPLE = "D9D2E9"
LANE_COUNT = 2
PANEL_WIDTH = 11
PANEL_GUTTER = 1
SIDE_GUTTER = 1
RAID_COLUMNS = 5
RAID_ROWS = 3

thin_black = Side(style="thin", color=BLACK)
medium_black = Side(style="medium", color=BLACK)


def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def raid_border(row_offset, column_offset):
    return Border(
        left=medium_black if column_offset == 0 else thin_black,
        right=medium_black if column_offset == RAID_COLUMNS - 1 else thin_black,
        top=medium_black if row_offset == 0 else thin_black,
        bottom=medium_black if row_offset == RAID_ROWS - 1 else thin_black,
    )


def member_display_name(member):
    name = str(member.get("name") or "").strip()
    label = str(member.get("label") or member.get("tier") or "").strip()
    return f"{name}-{label}" if label else name


def member_role_label(member):
    return str(member.get("label") or member.get("tier") or member.get("role") or "").strip()


def ordered_members(raid):
    members = list(raid.get("members") or [])
    return sorted(
        enumerate(members),
        key=lambda item: (item[1].get("role") == "Support", item[0]),
    )


def write_raid_block(sheet, start_row, start_column, raid):
    color_hex = COLOR_HEX.get(raid.get("color"), "D9D9D9")
    label_values = [
        f"{raid.get('name', '')} {raid.get('difficulty', '')}".strip(),
        "Reclear",
        "DONE" if raid.get("status") == "DONE" else "",
    ]

    for row_offset, value in enumerate(label_values):
        cell = sheet.cell(start_row + row_offset, start_column, value)
        cell.fill = fill(color_hex)
        cell.font = Font(name="Arial", size=8, bold=row_offset == 0, color=BLACK)
        cell.border = raid_border(row_offset, 0)

    members = [member for _, member in ordered_members(raid)][:4]
    while len(members) < 4:
        members.append({})

    for member_offset, member in enumerate(members, start=1):
        role_fill = SUPPORT_PURPLE if member.get("role") == "Support" else DPS_GREEN
        values = [
            member_display_name(member),
            member_role_label(member),
            member.get("itemLevel") if member.get("itemLevel") is not None else "",
        ]
        fills = [HEADER_BLUE, role_fill, role_fill]

        for row_offset, value in enumerate(values):
            cell = sheet.cell(start_row + row_offset, start_column + member_offset, value)
            cell.fill = fill(fills[row_offset])
            cell.font = Font(name="Arial", size=8, color=BLACK)
            cell.alignment = Alignment(
                horizontal="left" if row_offset == 0 else "right",
                vertical="center",
            )
            cell.border = raid_border(row_offset, member_offset)


def cluster_height(cluster):
    serca_count = sum(1 for raid in cluster if raid.get("name") == "Serca")
    cathedral_count = sum(1 for raid in cluster if raid.get("name") == "Cathedral")
    return max(serca_count, cathedral_count, 1) * RAID_ROWS + 1


def write_cluster(sheet, start_row, start_column, cluster):
    sercas = [raid for raid in cluster if raid.get("name") == "Serca"]
    cathedrals = [raid for raid in cluster if raid.get("name") == "Cathedral"]
    pair_count = max(len(sercas), len(cathedrals), 1)
    cathedral_column = start_column + RAID_COLUMNS + SIDE_GUTTER

    for index in range(pair_count):
        raid_row = start_row + index * RAID_ROWS
        if index < len(sercas):
            write_raid_block(sheet, raid_row, start_column, sercas[index])
        if index < len(cathedrals):
            write_raid_block(sheet, raid_row, cathedral_column, cathedrals[index])

    separator_row = start_row + pair_count * RAID_ROWS
    sheet.row_dimensions[separator_row].height = 9


def get_clusters(suggestion):
    clusters = suggestion.get("clusters")
    if clusters:
        return clusters

    grouped = {}
    for raid in suggestion.get("raids") or []:
        grouped.setdefault(raid.get("color") or "Unknown", []).append(raid)
    return list(grouped.values())


def configure_sheet(sheet):
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.freeze_panes = None

    for lane in range(LANE_COUNT):
        panel_start = 1 + lane * (PANEL_WIDTH + PANEL_GUTTER)
        sheet.column_dimensions[sheet.cell(1, panel_start).column_letter].width = 14
        for offset in range(1, RAID_COLUMNS):
            sheet.column_dimensions[sheet.cell(1, panel_start + offset).column_letter].width = 17
        sheet.column_dimensions[sheet.cell(1, panel_start + RAID_COLUMNS).column_letter].width = 2
        cathedral_start = panel_start + RAID_COLUMNS + SIDE_GUTTER
        sheet.column_dimensions[sheet.cell(1, cathedral_start).column_letter].width = 14
        for offset in range(1, RAID_COLUMNS):
            sheet.column_dimensions[sheet.cell(1, cathedral_start + offset).column_letter].width = 17

        if lane < LANE_COUNT - 1:
            gutter_column = panel_start + PANEL_WIDTH
            sheet.column_dimensions[sheet.cell(1, gutter_column).column_letter].width = 2


def build_option_sheet(workbook, suggestion, option_index):
    sheet = workbook.create_sheet(f"Option {option_index}")
    configure_sheet(sheet)
    lane_rows = [1] * LANE_COUNT

    for cluster in get_clusters(suggestion):
        lane = min(range(LANE_COUNT), key=lambda index: lane_rows[index])
        start_column = 1 + lane * (PANEL_WIDTH + PANEL_GUTTER)
        write_cluster(sheet, lane_rows[lane], start_column, cluster)
        lane_rows[lane] += cluster_height(cluster)

    max_row = max(lane_rows)
    max_column = LANE_COUNT * PANEL_WIDTH + (LANE_COUNT - 1) * PANEL_GUTTER
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            if cell.value is None and cell.fill.fill_type is None:
                cell.fill = fill(BLACK)

    sheet.sheet_properties.tabColor = "D95F43"
    return sheet


def generate_artifacts(input_path, output_directory):
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    suggestions = payload.get("suggestions") or []
    if not suggestions:
        raise ValueError("No optimizer suggestions were provided.")

    workbook = Workbook()
    workbook.remove(workbook.active)

    for option_index, suggestion in enumerate(suggestions, start=1):
        build_option_sheet(workbook, suggestion, option_index)

    output_directory.mkdir(parents=True, exist_ok=True)
    workbook_path = output_directory / "raid-suggestions.xlsx"
    workbook.save(workbook_path)

    image_paths = []
    for option_index in range(1, len(suggestions) + 1):
        image_path = output_directory / f"raid-suggestion-option-{option_index}.png"
        render_schedule(
            workbook_path,
            image_path,
            f"Option {option_index}",
            max_width=4096,
        )
        image_paths.append(image_path)

    return workbook_path, image_paths


def main():
    parser = argparse.ArgumentParser(
        description="Generate an XLSX workbook and PNG preview for every raid optimizer suggestion."
    )
    parser.add_argument("input", type=Path, help="Optimizer result JSON")
    parser.add_argument("output", type=Path, help="Output directory")
    args = parser.parse_args()

    try:
        workbook_path, image_paths = generate_artifacts(args.input, args.output)
    except (OSError, ValueError, json.JSONDecodeError) as error:
        print(error)
        sys.exit(1)

    print(json.dumps({
        "workbook": str(workbook_path),
        "images": [str(path) for path in image_paths],
    }))


if __name__ == "__main__":
    main()
