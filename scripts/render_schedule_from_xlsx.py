import argparse
import math
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles.colors import COLOR_INDEX
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    print("Missing dependency. Run: python -m pip install -r scripts/requirements.txt")
    sys.exit(1)


DEFAULT_SHEET_NAME = "Copy of Serca+Cath"
DEFAULT_ROW_HEIGHT = 15.75
DEFAULT_COLUMN_WIDTH = 12.63
RENDER_SCALE = 4


def load_theme_colors(workbook):
    if not workbook.loaded_theme:
        return []

    namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    root = ET.fromstring(workbook.loaded_theme)
    color_scheme = root.find(".//a:clrScheme", namespace)
    colors_by_name = {}

    if color_scheme is None:
        return []

    for child in list(color_scheme):
        srgb = child.find(".//a:srgbClr", namespace)
        system = child.find(".//a:sysClr", namespace)
        value = srgb.attrib.get("val") if srgb is not None else None
        value = value or (system.attrib.get("lastClr") if system is not None else None)
        colors_by_name[child.tag.rsplit("}", 1)[-1]] = value if value else "000000"

    excel_theme_order = [
        "lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3",
        "accent4", "accent5", "accent6", "hlink", "folHlink",
    ]
    return [colors_by_name.get(name, "000000") for name in excel_theme_order]


def apply_tint(rgb, tint):
    if not tint:
        return rgb

    output = []
    for channel in rgb:
        if tint < 0:
            channel = channel * (1 + tint)
        else:
            channel = channel * (1 - tint) + 255 * tint
        output.append(max(0, min(255, round(channel))))
    return tuple(output)


def resolve_color(color, theme_colors, fallback=None):
    if color is None:
        return fallback

    value = None
    if color.type == "rgb" and color.rgb:
        value = color.rgb[-6:]
    elif color.type == "theme" and color.theme is not None and color.theme < len(theme_colors):
        value = theme_colors[color.theme]
    elif color.type == "indexed" and color.indexed is not None and color.indexed < len(COLOR_INDEX):
        value = COLOR_INDEX[color.indexed][-6:]

    if not value:
        return fallback

    try:
        rgb = tuple(int(value[index:index + 2], 16) for index in (0, 2, 4))
    except (TypeError, ValueError):
        return fallback

    return apply_tint(rgb, color.tint)


def column_width_pixels(width):
    width = DEFAULT_COLUMN_WIDTH if width is None else width
    return max(1, math.floor(((256 * width + math.floor(128 / 7)) / 256) * 7))


def row_height_pixels(height):
    height = DEFAULT_ROW_HEIGHT if height is None else height
    return max(1, round(height))


def find_font(font_name, bold, italic, size):
    candidates = []
    suffix = ""
    if bold and italic:
        suffix = "bi"
    elif bold:
        suffix = "bd"
    elif italic:
        suffix = "i"

    if font_name and font_name.lower() == "arial":
        candidates.extend([
            Path(f"C:/Windows/Fonts/arial{suffix}.ttf"),
            Path("/usr/share/fonts/truetype/msttcorefonts/Arial_Bold_Italic.ttf" if bold and italic else
                 "/usr/share/fonts/truetype/msttcorefonts/Arial_Bold.ttf" if bold else
                 "/usr/share/fonts/truetype/msttcorefonts/Arial_Italic.ttf" if italic else
                 "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf"),
        ])

    dejavu_name = "DejaVuSans"
    if bold:
        dejavu_name += "-Bold"
    if italic:
        dejavu_name += "Oblique"
    candidates.append(Path(f"/usr/share/fonts/truetype/dejavu/{dejavu_name}.ttf"))

    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)

    return ImageFont.load_default()


def border_width(style):
    if style in {"medium", "mediumDashDot", "mediumDashDotDot", "mediumDashed"}:
        return 2 * RENDER_SCALE
    if style in {"thick", "double"}:
        return 3 * RENDER_SCALE
    return RENDER_SCALE if style else 0


def draw_cell_border(draw, cell, bounds, theme_colors):
    left, top, right, bottom = bounds
    sides = {
        "left": ((left, top, left, bottom), cell.border.left),
        "right": ((right, top, right, bottom), cell.border.right),
        "top": ((left, top, right, top), cell.border.top),
        "bottom": ((left, bottom, right, bottom), cell.border.bottom),
    }

    for coordinates, side in sides.values():
        if side is None or not side.style:
            continue
        color = resolve_color(side.color, theme_colors, (35, 35, 35))
        draw.line(coordinates, fill=color, width=border_width(side.style))


def draw_cell_text(image, cell, bounds, theme_colors, has_dropdown):
    if cell.value is None:
        return

    left, top, right, bottom = bounds
    width = max(1, right - left)
    height = max(1, bottom - top)
    layer = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    font_size = max(8, round((cell.font.sz or 10) * 0.8 * RENDER_SCALE))
    font = find_font(cell.font.name, cell.font.bold, cell.font.italic, font_size)
    color = resolve_color(cell.font.color, theme_colors, (0, 0, 0))
    if isinstance(cell.value, float) and cell.value.is_integer():
        text = str(int(cell.value))
    else:
        text = str(cell.value)
    padding = 3 * RENDER_SCALE
    dropdown_space = 8 * RENDER_SCALE if has_dropdown else 0
    text_box = draw.textbbox((0, 0), text, font=font)
    text_width = text_box[2] - text_box[0]
    text_height = text_box[3] - text_box[1]

    horizontal = cell.alignment.horizontal
    if horizontal == "center":
        x = (width - text_width) / 2
    elif horizontal == "right" or (horizontal is None and isinstance(cell.value, (int, float))):
        x = width - text_width - padding - dropdown_space
    else:
        x = padding

    vertical = cell.alignment.vertical
    if vertical == "top":
        y = padding - text_box[1]
    elif vertical == "bottom":
        y = height - text_height - padding - text_box[1]
    else:
        y = (height - text_height) / 2 - text_box[1]

    draw.text((x, y), text, font=font, fill=(*color, 255))

    if has_dropdown:
        arrow_x = width - 5 * RENDER_SCALE
        arrow_y = height // 2
        draw.polygon(
            [
                (arrow_x - 3 * RENDER_SCALE, arrow_y - RENDER_SCALE),
                (arrow_x + 3 * RENDER_SCALE, arrow_y - RENDER_SCALE),
                (arrow_x, arrow_y + 2 * RENDER_SCALE),
            ],
            fill=(46, 95, 130, 255),
        )

    image.alpha_composite(layer, (left, top))


def get_dropdown_cells(sheet):
    cells = set()
    for validation in sheet.data_validations.dataValidation:
        if validation.type != "list" or validation.showDropDown is True:
            continue
        for cell_range in validation.ranges.ranges:
            for row in sheet.iter_rows(
                min_row=cell_range.min_row,
                max_row=cell_range.max_row,
                min_col=cell_range.min_col,
                max_col=cell_range.max_col,
            ):
                cells.update(cell.coordinate for cell in row)
    return cells


def render_schedule(workbook_path, output_path, sheet_name, max_width):
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet not found: {sheet_name}. Available sheets: {available}")

    sheet = workbook[sheet_name]
    theme_colors = load_theme_colors(workbook)
    dropdown_cells = get_dropdown_cells(sheet)
    populated_cells = [cell for row in sheet.iter_rows() for cell in row if cell.value is not None]
    if not populated_cells:
        raise ValueError(f"Sheet is empty: {sheet_name}")
    first_populated_row = min(cell.row for cell in populated_cells)
    last_populated_row = max(cell.row for cell in populated_cells)
    first_populated_column = min(cell.column for cell in populated_cells)
    last_populated_column = max(cell.column for cell in populated_cells)
    column_widths = [
        column_width_pixels(sheet.column_dimensions[sheet.cell(1, column).column_letter].width)
        * RENDER_SCALE
        for column in range(1, sheet.max_column + 1)
    ]
    row_heights = [
        row_height_pixels(sheet.row_dimensions[row].height) * RENDER_SCALE
        for row in range(1, sheet.max_row + 1)
    ]

    x_positions = [0]
    for width in column_widths:
        x_positions.append(x_positions[-1] + width)
    y_positions = [0]
    for height in row_heights:
        y_positions.append(y_positions[-1] + height)

    image = Image.new("RGBA", (x_positions[-1], y_positions[-1]), (0, 0, 0, 255))
    draw = ImageDraw.Draw(image)

    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            bounds = (
                x_positions[column - 1],
                y_positions[row - 1],
                x_positions[column],
                y_positions[row],
            )
            fill = None
            if cell.fill.fill_type == "solid":
                fill = resolve_color(cell.fill.fgColor, theme_colors)
            if fill is not None:
                draw.rectangle(bounds, fill=fill)

    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            bounds = (
                x_positions[column - 1],
                y_positions[row - 1],
                x_positions[column],
                y_positions[row],
            )
            draw_cell_border(draw, cell, bounds, theme_colors)
            draw_cell_text(image, cell, bounds, theme_colors, cell.coordinate in dropdown_cells)

    crop_margin = 8 * RENDER_SCALE
    image = image.crop((
        max(0, x_positions[first_populated_column - 1] - crop_margin),
        max(0, y_positions[first_populated_row - 1] - crop_margin),
        min(image.width, x_positions[last_populated_column] + crop_margin),
        min(image.height, y_positions[last_populated_row] + crop_margin),
    ))

    if max_width and image.width > max_width:
        target_height = round(image.height * max_width / image.width)
        image = image.resize((max_width, target_height), Image.Resampling.LANCZOS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.convert("RGB").save(output_path, "PNG", optimize=True)


def main():
    global RENDER_SCALE

    parser = argparse.ArgumentParser(description="Render a raid schedule worksheet as a PNG image.")
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx workbook")
    parser.add_argument("output", type=Path, help="Path for the generated .png")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Worksheet to render")
    parser.add_argument("--max-width", default=4096, type=int, help="Maximum output width in pixels")
    parser.add_argument("--render-scale", default=4, type=int, help="Internal supersampling scale")
    args = parser.parse_args()
    RENDER_SCALE = max(1, args.render_scale)

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}")
        sys.exit(1)

    try:
        render_schedule(args.workbook, args.output, args.sheet, args.max_width)
    except (OSError, ValueError) as error:
        print(error)
        sys.exit(1)

    print(f"Rendered {args.sheet} to {args.output}")


if __name__ == "__main__":
    main()
